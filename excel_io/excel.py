import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import re
from models import MedecinData

# Constantes
_NbMedecins = 17
_NbCrenau = 55
_Gmax= 100

def sanitize(s):
    return re.sub(r'[^a-zA-Z0-9]', '', s)

def get_lines(df):
    return df.iloc[list(range(3, 3 + _NbCrenau))]

def get_creneaux(df):
    # Generate the list of dates from column D (index 3)
    return df.iloc[3:3 + _NbCrenau, 3].tolist()

class Voeux: 
    def __init__(self, garde_file_path, astreinte_file_path):
        self.garde_file_path = garde_file_path
        self.astreinte_file_path = astreinte_file_path

        df_g = pd.read_excel(garde_file_path, sheet_name=0)
        df_g = df_g.map(lambda x: sanitize(x) if isinstance(x, str) else x)

        df_a = pd.read_excel(astreinte_file_path, sheet_name=0)
        df_a = df_a.map(lambda x: sanitize(x) if isinstance(x, str) else x)
        
        dict_of_columns = {
            df_g.iloc[1, col]: [(df_g.iloc[i, col], df_a.iloc[i, col]) for i in range(3, 3 + _NbCrenau)] for col in range(9, 9 + _NbMedecins)
        }

        self.creneaux = get_creneaux(df_g)
        self.data = dict_of_columns

    def adjust_time(self, begin, end):
        # get creneaux id from begin to end
        begin_id = self.creneaux.index(begin)
        end_id = self.creneaux.index(end)

        # adjust data
        self.data = {medecin: self.data[medecin][begin_id:end_id + 1] for medecin in self.data}
        return begin_id, end_id
    
    def __str__(self):
        return f"Voeux[{self.garde_file_path, self.astreinte_file_path}]: {self.data}"


class ExcelSolution:
    def __init__(self, file_path):
        self.file_path = file_path
    
    def read_solution(self):
        df = pd.read_excel(self.file_path, sheet_name=0)
        df = df.map(lambda x: sanitize(x) if isinstance(x, str) else None)
        lines = get_lines(df)
        self.solution = list(lines.iloc[:, [4, 5]].itertuples(index=False, name=None))
        return self.solution
    
    def write_solution(self, solution, file_path, begin_index):
        wb = load_workbook(self.file_path)
        ws = wb.active

        for i, creneau in enumerate(solution):
            ws.cell(row=5 + begin_index + i, column=5).value = creneau[0]
            ws.cell(row=5 + begin_index + i, column=6).value = creneau[1]

        wb.save(file_path)

    def write_graph(self, satisfaction, file_path):
        wb = load_workbook(file_path)

        graph_sheet = wb.create_sheet(title="Satisfaction")
        graph_sheet.cell(row=1, column=1, value="Medecin")
        graph_sheet.cell(row=1, column=2, value="Satisfaction")

        for i, (medecin, score) in enumerate(satisfaction.items(), start=2):
            graph_sheet.cell(row=i, column=1, value=medecin)
            graph_sheet.cell(row=i, column=2, value=score)

        chart = BarChart()
        chart.title = "Satisfaction par Médecin"
        chart.y_axis.title = "Satisfaction"
        chart.x_axis.title = "Médecin"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        data = Reference(graph_sheet, min_col=2, min_row=1, max_row=1 + len(satisfaction))
        categories = Reference(graph_sheet, min_col=1, min_row=2, max_row=1 + len(satisfaction))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        graph_sheet.add_chart(chart, "E5")
        wb.save(file_path)

    def compute_blocked_gardes(self):
        return {i: sol[0] for i, sol in enumerate(self.solution) if sol[0] != None}
    
    def compute_blocked_astreintes(self):
        return {i: sol[1] for i, sol in enumerate(self.solution) if sol[1] != None}


class ExcelMedecin:
    def __init__(self, file_path):
        self.file_path = file_path

    def read_medecins(self):
        df = pd.read_excel(self.file_path, sheet_name=0)
        self.medecins = df.iloc[:, 0].tolist()
        self.medecins_garde = df[df.iloc[:, 1]].iloc[:, 0].tolist()
        self.medecins_astreinte = df[df.iloc[:, 2]].iloc[:, 0].tolist()
        self.specialisation = df[df.iloc[:, 3]].iloc[:, 0].tolist()
        self.voeux_nbr_gardes = {medecin: df[df.iloc[:, 0] == medecin].iloc[:, 4].tolist()[0] for medecin in self.medecins}
        return MedecinData(self.medecins_garde, self.medecins_astreinte, self.specialisation, _Gmax, self.voeux_nbr_gardes)
